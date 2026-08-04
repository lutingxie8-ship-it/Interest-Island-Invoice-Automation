# src/report_generator.py
# 职责：将处理后数据生成为 .md（给大模型）和 .xlsx（给人看）双报告

from dataclasses import dataclass, field
from datetime import datetime
import os
from typing import Optional


# ── ReportData 聚合数据结构 ──

@dataclass
class ReportData:
    """报告生成所需的全部已聚合数据"""
    # 统计
    run_time: str               # 运行时间字符串，如 "2026-07-29 17:39:53"
    total_unread: int           # 处理未读邮件数
    invoice_count: int          # 开票邮件数
    urgent_count: int           # 加急订单数
    uncertain_count: int        # 疑似邮件数
    failed_count: int           # 解析失败数

    # 去重后的有效订单（已按分类分组）
    urgent_orders: list         # 加急订单（urgent_valid + urgent_invalid）
    normal_orders: list         # 正常订单（normal_valid）
    invalid_orders: list        # 异常订单（urgent_invalid + normal_invalid）

    # 疑似不确定邮件
    uncertain_entries: list     # list[dict] 含 classification + message

    # 解析失败明细
    failed_entries: list        # list[dict] 含 subject/sender/date/message_id/reason


# ── 入口 ──

def generate_report(data: 'ReportData', output_dir: str) -> tuple[str, str]:
    """
    入口：生成 .md 和 .xlsx 双报告。

    Args:
        data: 已聚合的 ReportData（由调用方构建）
        output_dir: 输出目录路径（自动创建）

    Returns:
        (md_path, xlsx_path) 双报告文件完整路径
    """
    # 1. 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)

    # 2. 生成文件名前缀
    now = datetime.now()
    filename_prefix = (
        f"{now.year}.{now.month}.{now.day}_"
        f"{now.hour:02d}-{now.minute:02d}_发票邮件报告"
    )

    # 3. 生成 .md
    md_path = os.path.join(output_dir, filename_prefix + ".md")
    _generate_md(data, md_path)

    # 4. 生成 .xlsx
    xlsx_path = os.path.join(output_dir, filename_prefix + ".xlsx")
    _generate_xlsx(data, xlsx_path)

    return (md_path, xlsx_path)


# ── 数据聚合 ──
# 注：ReportData 由各调用方（parse.py）直接构造，这里不再提供聚合函数。

# ════════════════════════════════════════════
# Wave 2: MD 报告生成
# ════════════════════════════════════════════

def _generate_md(data: 'ReportData', output_path: str):
    """生成 .md 报告文件。"""
    lines = []

    # ── 头部：运行时间 & 统计摘要 ──
    lines.append("# 发票邮件处理报告")
    lines.append("")
    lines.append(f"**运行时间**：{data.run_time}")
    lines.append("")
    lines.append("## 处理统计")
    lines.append("")
    lines.append(f"- 本次处理未读邮件：**{data.total_unread}** 封")
    lines.append(f"- 其中开票邮件：**{data.invoice_count}** 封")
    lines.append(f"- 加急订单：**{data.urgent_count}** 笔")
    lines.append(f"- 异常订单：**{len(data.invalid_orders)}** 笔")
    lines.append(f"- 疑似不确定邮件：**{data.uncertain_count}** 封")
    lines.append(f"- 解析失败：**{data.failed_count}** 笔")
    lines.append("")

    # ── 加急订单区（最前，订单号加粗）──
    if data.urgent_orders:
        lines.append("## 加急订单")
        lines.append("")
        lines.append("| 订单号 | 开票金额 | 备注 | 发票抬头 | 税号 | 来源邮件主题 | 发件人 | 邮件时间 |")
        lines.append("|---|---|---|---|---|---|---|---|")
        for o in data.urgent_orders:
            order_id_bold = f"**{o.order_id_cleaned}**"
            lines.append(
                f"| {order_id_bold} | {o.amount_raw} | {o.note} | {o.title} | {o.tax_id} | "
                f"{o.message_subject} | {o.message_sender} | {o.message_date} |"
            )
        lines.append("")

    # ── 正常订单区 ──
    if data.normal_orders:
        lines.append("## 正常订单")
        lines.append("")
        lines.append("| 订单号 | 开票金额 | 备注 | 发票抬头 | 税号 | 来源邮件主题 | 发件人 | 邮件时间 |")
        lines.append("|---|---|---|---|---|---|---|---|")
        for o in data.normal_orders:
            lines.append(
                f"| {o.order_id_cleaned} | {o.amount_raw} | {o.note} | {o.title} | {o.tax_id} | "
                f"{o.message_subject} | {o.message_sender} | {o.message_date} |"
            )
        lines.append("")

    # ── 订单号异常区 ──
    if data.invalid_orders:
        lines.append("## 订单号异常")
        lines.append("")
        lines.append("| 订单号原文 | 清洗后数字 | 异常原因 | 开票金额 | 备注 | 来源邮件主题 | 发件人 | 邮件时间 |")
        lines.append("|---|---|---|---|---|---|---|---|")
        for o in data.invalid_orders:
            lines.append(
                f"| {o.order_id_original} | {o.order_id_cleaned} | {o.validation_reason} | "
                f"{o.amount_raw} | {o.note} | "
                f"{o.message_subject} | {o.message_sender} | {o.message_date} |"
            )
        lines.append("")

    # ── 疑似不确定邮件区 ──
    if data.uncertain_entries:
        lines.append("## 疑似不确定邮件")
        lines.append("")
        lines.append("| 邮件主题 | 发件人 | 邮件时间 | 不确定原因 |")
        lines.append("|---|---|---|---|")
        for entry in data.uncertain_entries:
            msg = entry.get("message")
            cls = entry.get("classification")
            reasons = "; ".join(cls.reasons) if cls else ""
            subject = msg.subject if msg else ""
            sender = msg.sender if msg else ""
            date = msg.date if msg else ""
            lines.append(
                f"| {subject} | {sender} | {date} | {reasons} |"
            )
        lines.append("")

    # ── 解析失败区 ──
    if data.failed_entries:
        lines.append("## 解析失败")
        lines.append("")
        lines.append("| 邮件主题 | 发件人 | 邮件时间 | 失败原因 |")
        lines.append("|---|---|---|---|")
        for entry in data.failed_entries:
            lines.append(
                f"| {entry.get('subject', '')} | {entry.get('sender', '')} | "
                f"{entry.get('date', '')} | {entry.get('reason', '')} |"
            )
        lines.append("")

    # 写入文件（UTF-8）
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ════════════════════════════════════════════
# Wave 3: XLSX 报告生成
# ════════════════════════════════════════════

def _generate_xlsx(data: 'ReportData', output_path: str):
    """
    生成 .xlsx 报告文件（5 个 Sheet）。

    Sheet 1: 汇总 —— 运行时间、各分类数量
    Sheet 2: 加急订单 —— 订单号加粗、整行标黄
    Sheet 3: 正常订单
    Sheet 4: 订单号异常
    Sheet 5: 疑似不确定邮件
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # ── Sheet 1: 汇总 ──
    ws1 = wb.active
    ws1.title = "汇总"
    _write_summary_sheet(ws1, data)

    # ── Sheet 2: 加急订单 ──
    ws2 = wb.create_sheet("加急订单")
    _write_order_sheet(ws2, data.urgent_orders, is_urgent=True)

    # ── Sheet 3: 正常订单 ──
    ws3 = wb.create_sheet("正常订单")
    _write_order_sheet(ws3, data.normal_orders, is_urgent=False)

    # ── Sheet 4: 订单号异常 ──
    ws4 = wb.create_sheet("订单号异常")
    _write_invalid_sheet(ws4, data.invalid_orders)

    # ── Sheet 5: 疑似不确定邮件 ──
    ws5 = wb.create_sheet("疑似不确定邮件")
    _write_uncertain_sheet(ws5, data.uncertain_entries)

    # ── Sheet 6: 解析失败 ──
    ws6 = wb.create_sheet("解析失败")
    _write_failed_sheet(ws6, data.failed_entries)

    wb.save(output_path)


def _write_summary_sheet(ws, data: 'ReportData'):
    """写入汇总 Sheet。"""
    from openpyxl.styles import Font
    headers = ["项目", "数值"]
    ws.append(headers)
    ws.append(["运行时间", data.run_time])
    ws.append(["处理未读邮件数", data.total_unread])
    ws.append(["开票邮件数", data.invoice_count])
    ws.append(["加急订单数", data.urgent_count])
    ws.append(["正常订单数", len(data.normal_orders)])
    ws.append(["异常订单数", len(data.invalid_orders)])
    ws.append(["疑似不确定邮件数", data.uncertain_count])
    ws.append(["解析失败数", data.failed_count])

    # 表头加粗
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _write_order_sheet(ws, orders, is_urgent: bool):
    """
    写入订单 Sheet（加急 / 正常）。

    列：订单号 | 开票金额 | 备注 | 来源邮件主题 | 发件人 | 邮件时间
    """
    from openpyxl.styles import Font, PatternFill
    headers = ["订单号", "开票金额", "备注", "发票抬头", "税号", "来源邮件主题", "发件人", "邮件时间"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # 加急 Sheet 样式
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)

    for order in orders:
        row_data = [
            order.order_id_cleaned,
            order.amount_raw,
            order.note,
            order.title,
            order.tax_id,
            order.message_subject,
            order.message_sender,
            order.message_date,
        ]
        ws.append(row_data)

        if is_urgent:
            # 加急：订单号加粗 + 整行标黄
            row_idx = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = yellow_fill
            # 订单号列（第 1 列）额外加粗
            ws.cell(row=row_idx, column=1).font = bold_font


def _write_invalid_sheet(ws, orders):
    """
    写入异常 Sheet。

    列：订单号原文 | 清洗后数字 | 异常原因 | 开票金额 | 备注 | 来源邮件主题 | 发件人 | 邮件时间
    """
    from openpyxl.styles import Font
    headers = ["订单号原文", "清洗后数字", "异常原因", "开票金额", "备注", "来源邮件主题", "发件人", "邮件时间"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for order in orders:
        ws.append([
            order.order_id_original,
            order.order_id_cleaned,
            order.validation_reason,
            order.amount_raw,
            order.note,
            order.message_subject,
            order.message_sender,
            order.message_date,
        ])


def _write_uncertain_sheet(ws, entries):
    """
    写入疑似不确定邮件 Sheet。

    列：邮件主题 | 发件人 | 邮件时间 | 不确定原因
    """
    from openpyxl.styles import Font
    headers = ["邮件主题", "发件人", "邮件时间", "不确定原因"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for entry in entries:
        msg = entry.get("message")
        cls = entry.get("classification")
        reasons = "; ".join(cls.reasons) if cls else ""
        ws.append([
            msg.subject if msg else "",
            msg.sender if msg else "",
            msg.date if msg else "",
            reasons,
        ])


def _write_failed_sheet(ws, entries):
    """
    写入解析失败 Sheet。

    列：邮件主题 | 发件人 | 邮件时间 | 失败原因
    """
    from openpyxl.styles import Font
    headers = ["邮件主题", "发件人", "邮件时间", "失败原因"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for entry in entries:
        ws.append([
            entry.get("subject", ""),
            entry.get("sender", ""),
            entry.get("date", ""),
            entry.get("reason", ""),
        ])
