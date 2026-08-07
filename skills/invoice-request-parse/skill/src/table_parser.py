# src/table_parser.py
# 职责：从邮件 HTML 正文中解析《开票申请汇总表》，提取 3 个核心字段

from dataclasses import dataclass
from html.parser import HTMLParser

from skill.src.logger import setup_logger

logger = setup_logger("table_parser")


# ──────────────────────────────────────────────
# Data classes
# ──────────────────────────────────────────────


@dataclass
class ParsedOrder:
    amount_raw: str            # 开票金额原文（如 "3904元"、"1880元"）
    order_id_raw: str          # 订单号原文（可能含前缀，如 "主订单ID：9000000000000001"）
    note: str                  # 备注原文
    title: str = ""            # 发票抬头原文（如 "广州沃道投资管理有限公司"）— 供下游开票
    tax_id: str = ""           # 税号/统一社会信用代码 — 供下游开票（企业必填）


@dataclass
class TableParseResult:
    orders: list[ParsedOrder]  # 解析成功的订单列表（多行 = 多条）
    raw_table: str | None      # 原始表格 HTML（用于调试/日志）
    success: bool              # True = 至少成功解析 1 行


# ──────────────────────────────────────────────
# HTML Table Extractor（基于 html.parser）
# ──────────────────────────────────────────────


class HTMLTableExtractor(HTMLParser):
    """从 HTML 中提取所有顶层 <table> 的原始 HTML（嵌套表格归入其父表）"""

    def __init__(self):
        super().__init__()
        self._tables: list[str] = []
        self._depth = 0
        self._current_table: list[str] | None = None

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._depth += 1
            if self._depth == 1:
                # 进入一个新的顶层表格，开始捕获
                self._current_table = ["<table>"]
            elif self._current_table is not None:
                # 嵌套表格，原样记录其开始标签
                self._current_table.append(self.get_starttag_text())
        elif self._current_table is not None:
            self._current_table.append(self.get_starttag_text())

    def handle_endtag(self, tag):
        if tag == "table":
            if self._current_table is not None:
                self._current_table.append("</table>")
                self._depth -= 1
                if self._depth == 0:
                    # 顶层表格闭合，输出捕获结果
                    self._tables.append("".join(self._current_table))
                    self._current_table = None
        elif self._current_table is not None:
            self._current_table.append(f"</{tag}>")

    def handle_data(self, data):
        if self._current_table is not None:
            self._current_table.append(data)

    def handle_entityref(self, name):
        """保留 HTML 实体（如 &nbsp;）"""
        if self._current_table is not None:
            self._current_table.append(f"&{name};")

    def handle_charref(self, name):
        """保留字符引用（如 &#160;）"""
        if self._current_table is not None:
            self._current_table.append(f"&#{name};")

    def get_tables(self) -> list[str]:
        return self._tables


# ──────────────────────────────────────────────
# TableParser
# ──────────────────────────────────────────────


class TableParser:
    def __init__(self, config: dict):
        """
        从 config 读取字段映射：
        - table_parser.mode_priority: 优先尝试的模式列表 ["key_value", "column_index"]
        - table_parser.field_mapping: 键值对模式的字段名映射
          amount: ["开票金额", "金额", "实付金额"]
          order_id: ["订单号", "订单编号", "主订单ID"]
          note: ["备注"]
        - table_parser.column_indices: 列索引模式的列号配置
          amount: 8, order_id: 10, note: 14
        - keywords.invoice_table: 表格标题关键词（默认 "发票申请"）
        """
        table_cfg = config.get("keywords", {}).get("table_parser", {})
        default_mapping = {
            "amount": ["开票金额", "金额", "实付金额"],
            "order_id": ["订单号", "订单编号", "主订单ID"],
            "note": ["备注"],
            "title": ["发票抬头", "抬头", "购买方名称", "公司名称"],
            "tax_id": ["税号", "统一社会信用代码", "纳税人识别号", "购方识别号"],
        }
        default_indices = {"amount": 8, "order_id": 10, "note": 14}

        self._mode_priority = table_cfg.get("mode_priority", ["key_value", "column_index"])
        self._field_mapping = table_cfg.get("field_mapping", default_mapping)
        self._column_indices = table_cfg.get("column_indices", default_indices)
        self._table_keyword = config.get("keywords", {}).get("invoice_table", "发票申请")

    # ── 公开方法 ──

    def parse(self, body_html: str | None) -> TableParseResult:
        """
        从 body_html 中解析表格。

        流程：
        1. 如果 body_html 为 None 或空字符串 → 返回空结果
        2. 用 HTMLTableExtractor 提取所有 table
        3. 对每个 table，检查是否含 self._table_keyword
        4. 对匹配的 table 按 mode_priority 顺序尝试解析
        5. 返回 TableParseResult
        """
        # Step 1: 空输入检查
        if not body_html or not body_html.strip():
            return TableParseResult(orders=[], raw_table=None, success=False)

        # Step 2: 提取所有表格
        extractor = HTMLTableExtractor()
        extractor.feed(body_html)
        tables = extractor.get_tables()

        if not tables:
            return TableParseResult(orders=[], raw_table=None, success=False)

        # Step 3-4: 按关键词匹配并尝试解析
        for table_html in tables:
            # 检查关键词
            if self._table_keyword not in table_html:
                continue

            orders: list[ParsedOrder] | None = None

            # 按优先级依次尝试解析模式
            for mode in self._mode_priority:
                if mode == "key_value":
                    orders = self._parse_key_value(table_html)
                elif mode == "column_index":
                    orders = self._parse_column_index(table_html)

                if orders is not None:
                    break

            # 解析成功 → 聚合多行并返回
            if orders is not None:
                aggregated = self._multi_row_aggregate(orders)
                return TableParseResult(
                    orders=aggregated,
                    raw_table=table_html,
                    success=True,
                )

        # 所有表格均未匹配或解析失败
        return TableParseResult(orders=[], raw_table=None, success=False)

    # ── 公开方法：xlsx 直读（优化 A：跳过 HTML 中转）──

    def parse_xlsx(self, xlsx_bytes: bytes | None) -> TableParseResult:
        """
        直接解析 .xlsx 附件字节（openpyxl 直读，不经过 HTML 转换）。

        流程：
        1. 空输入检查
        2. openpyxl 打开工作簿（data_only=True 取计算值）
        3. 选表：优先 sheet 名含 invoice_table 关键字；否则第一个非空 sheet
        4. 读全部行 → 按表头别名定位字段列（比写死列号更准）
        5. 宽表走列索引模式、窄表（≤3 列）走键值对模式
        6. 聚合多行并返回
        """
        if not xlsx_bytes:
            return TableParseResult(orders=[], raw_table=None, success=False)

        try:
            import io
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("openpyxl 未安装，无法直读 .xlsx 附件")
            return TableParseResult(orders=[], raw_table=None, success=False)

        try:
            wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        except Exception as e:
            logger.warning(f".xlsx 打开失败: {e}")
            return TableParseResult(orders=[], raw_table=None, success=False)

        # 3. 选表
        target_ws = None
        for ws in wb.worksheets:
            if ws.title and self._table_keyword in ws.title:
                target_ws = ws
                break
        if target_ws is None:
            for ws in wb.worksheets:
                if ws.max_row and ws.max_column:
                    target_ws = ws
                    break
        if target_ws is None:
            logger.warning("xlsx 中无可用工作表")
            return TableParseResult(orders=[], raw_table=None, success=False)

        # 4. 读全部行（单元格转字符串，去完全空行）
        rows: list[list[str]] = []
        for row in target_ws.iter_rows(values_only=True):
            rows.append([str(c).strip() if c is not None else "" for c in row])
        rows = [r for r in rows if any(cell.strip() for cell in r)]
        if not rows:
            return TableParseResult(orders=[], raw_table=None, success=False)

        # 5. 模式判定：窄表（≤3 列）走键值对，宽表走列索引
        max_col = max((len(r) for r in rows), default=0)
        if max_col <= 3:
            orders = self._parse_key_value_rows(rows)
        else:
            orders = self._parse_column_index_rows(rows)

        if not orders:
            logger.warning("xlsx 直读：未解析出任何订单")
            return TableParseResult(orders=[], raw_table=None, success=False)

        aggregated = self._multi_row_aggregate(orders)
        return TableParseResult(orders=aggregated, raw_table=None, success=True)

    # ── xlsx 内部解析器 ──

    def _find_header_row(self, rows: list[list[str]]) -> int | None:
        """扫描行，找表头行（命中字段别名最多的行），至少命中 2 个字段才认。"""
        best_idx = None
        best_score = 0
        for i, row in enumerate(rows):
            score = 0
            row_text = " ".join(row)
            for aliases in self._field_mapping.values():
                if any(alias in row_text for alias in aliases):
                    score += 1
            if score > best_score:
                best_score = score
                best_idx = i
        return best_idx if best_score >= 2 else None

    def _parse_column_index_rows(self, rows: list[list[str]]) -> list[ParsedOrder] | None:
        """
        宽表模式：基于表头别名定位字段列（不依赖写死列号）。
        找不到表头时回退到 config.column_indices（1-based）兜底。
        """
        header_idx = self._find_header_row(rows)
        if header_idx is None:
            logger.warning("宽表未找到表头行，尝试 column_indices 兜底")
            return self._parse_column_index_legacy(rows)

        header = rows[header_idx]
        # 每个字段只映射到「最佳匹配」的那一列（别名越长越优先、同长则列序靠前优先），
        # 避免「公司名称」与「商品名称」都含别名「名称」这类同名字段列冲突
        best_for_field: dict[str, tuple] = {}  # field -> (score, ci)
        for ci, cell in enumerate(header):
            for field, aliases in self._field_mapping.items():
                for alias in aliases:
                    if alias in cell:
                        score = (len(alias), -ci)  # 别名越长越优；同长时列越靠前越优
                        prev = best_for_field.get(field)
                        if prev is None or score > prev[0]:
                            best_for_field[field] = (score, ci)
        col_to_field = {ci: field for field, (_, ci) in best_for_field.items()}

        if "amount" not in col_to_field.values() and "order_id" not in col_to_field.values():
            logger.warning("表头未定位到 金额/订单号，尝试 column_indices 兜底")
            return self._parse_column_index_legacy(rows)

        orders: list[ParsedOrder] = []
        for row in rows[header_idx + 1:]:
            if not any(c.strip() for c in row):
                continue
            vals = {f: (row[ci].strip() if ci < len(row) else "") for ci, f in col_to_field.items()}
            amount = vals.get("amount", "")
            order_id = vals.get("order_id", "")
            if not amount and not order_id:
                continue
            orders.append(ParsedOrder(
                amount_raw=amount,
                order_id_raw=order_id,
                note=vals.get("note", ""),
                title=vals.get("title", ""),
                tax_id=vals.get("tax_id", ""),
            ))
        return orders if orders else None

    def _parse_column_index_legacy(self, rows: list[list[str]]) -> list[ParsedOrder] | None:
        """兜底：沿用 config.column_indices（1-based）按固定列号取数。"""
        indices = self._column_indices
        all_idx = [v for v in [indices.get("amount", 0), indices.get("order_id", 0),
                               indices.get("note", 0), indices.get("title", 0),
                               indices.get("tax_id", 0)] if v]
        if not all_idx:
            return None
        max_col = max(all_idx)
        orders: list[ParsedOrder] = []
        for row in rows:
            if len(row) <= max_col:
                continue
            amount = row[indices.get("amount", 8) - 1].strip()
            order_id = row[indices.get("order_id", 10) - 1].strip()
            note = row[indices.get("note", 14) - 1].strip()
            title = row[indices.get("title", 0) - 1].strip() if indices.get("title") else ""
            tax_id = row[indices.get("tax_id", 0) - 1].strip() if indices.get("tax_id") else ""
            if not amount and not order_id:
                continue
            orders.append(ParsedOrder(amount_raw=amount, order_id_raw=order_id,
                                       note=note, title=title, tax_id=tax_id))
        return orders if orders else None

    def _parse_key_value_rows(self, rows: list[list[str]]) -> list[ParsedOrder] | None:
        """键值对模式：2 列表格，按字段别名匹配 key。"""
        alias_to_field: dict[str, str] = {}
        for field, aliases in self._field_mapping.items():
            for alias in aliases:
                alias_to_field[alias] = field

        amount = order_id = note = title = tax_id = ""
        for row in rows:
            if len(row) < 2:
                continue
            matched = self._match_key(row[0], alias_to_field)
            val = row[1].strip()
            if matched == "amount":
                amount = val
            elif matched == "order_id":
                order_id = val
            elif matched == "note":
                note = val
            elif matched == "title":
                title = val
            elif matched == "tax_id":
                tax_id = val

        if not amount and not order_id:
            return None
        return [ParsedOrder(amount_raw=amount, order_id_raw=order_id,
                             note=note, title=title, tax_id=tax_id)]

    # ── 模式 A：键值对模式 ──

    def _parse_key_value(self, table_html: str) -> list[ParsedOrder] | None:
        """
        模式 A：2 列键值对模式
        提取所有 (key, value) 对，匹配配置的字段名映射。
        如果表格不匹配（无目标字段）返回 None。
        """
        parser = _KeyValueCellExtractor()
        parser.feed(table_html)
        pairs = parser.get_pairs()

        if not pairs:
            return None

        # 构建反向映射：任意别名 → 标准字段名
        alias_to_field: dict[str, str] = {}
        for field, aliases in self._field_mapping.items():
            for alias in aliases:
                alias_to_field[alias] = field

        # 从键值对中提取目标字段
        amount = ""
        order_id = ""
        note = ""
        title = ""
        tax_id = ""

        for key_text, value_text in pairs:
            matched_field = self._match_key(key_text, alias_to_field)
            if matched_field == "amount":
                amount = value_text
            elif matched_field == "order_id":
                order_id = value_text
            elif matched_field == "note":
                note = value_text
            elif matched_field == "title":
                title = value_text
            elif matched_field == "tax_id":
                tax_id = value_text

        # 必须至少匹配到金额或订单号其中之一
        if not amount and not order_id:
            return None

        return [ParsedOrder(amount_raw=amount, order_id_raw=order_id, note=note, title=title, tax_id=tax_id)]

    def _match_key(self, key_text: str, alias_to_field: dict[str, str]) -> str | None:
        """检查 key_text 是否包含任一别名，返回标准字段名或 None"""
        for alias, field in alias_to_field.items():
            if alias in key_text:
                return field
        return None

    # ── 模式 B：列索引宽表模式 ──

    def _parse_column_index(self, table_html: str) -> list[ParsedOrder] | None:
        """
        模式 B：15 列宽表模式
        第 1 行：大标题（跳过）
        第 2 行：表头（跳过）
        第 3 行起：数据行
        提取列索引 8（金额）、10（订单号）、14（备注）
        如果表格列数不足或行数不足返回 None。
        """
        parser = _WideTableCellExtractor()
        parser.feed(table_html)
        rows = parser.get_rows()

        # 至少需要 3 行（标题 + 表头 + 至少 1 行数据）
        if len(rows) < 3:
            return None

        indices = self._column_indices
        all_indices = [v for v in [indices.get("amount", 0), indices.get("order_id", 0), indices.get("note", 0), indices.get("title", 0), indices.get("tax_id", 0)] if v]
        max_col = max(all_indices) if all_indices else 0

        orders: list[ParsedOrder] = []

        # 从第 3 行（索引 2）开始读取数据行
        for row in rows[2:]:
            if len(row) <= max_col:
                return None  # 列数不足，该模式不匹配

            # column_indices 按 PRD 的 1-based 列号配置（第 8/10/14 列），
            # Python list 索引为 0-based，故统一 -1 转换
            amount = row[indices.get("amount", 8) - 1].strip()
            order_id = row[indices.get("order_id", 10) - 1].strip()
            note = row[indices.get("note", 14) - 1].strip()
            # title/tax_id 仅在配置了列号时提取（未配置则留空，不影响原逻辑）
            title = row[indices.get("title", 0) - 1].strip() if indices.get("title") else ""
            tax_id = row[indices.get("tax_id", 0) - 1].strip() if indices.get("tax_id") else ""

            orders.append(ParsedOrder(amount_raw=amount, order_id_raw=order_id, note=note, title=title, tax_id=tax_id))

        if not orders:
            return None

        return orders

    # ── 多行聚合 ──

    @staticmethod
    def _multi_row_aggregate(orders: list[ParsedOrder]) -> list[ParsedOrder]:
        """
        多行聚合：同一邮件多行 = 同一订单
        - 金额保留所有行明细（列表字符串，如 "3904元, 1880元"）
        - 订单号取第一个非空值
        - 备注取第一个非空值
        """
        if not orders:
            return []
        # Single order → return as-is
        if len(orders) == 1:
            return orders
        # Multiple rows → aggregate into one
        amounts = [o.amount_raw for o in orders if o.amount_raw.strip()]
        first_oid = next((o.order_id_raw for o in orders if o.order_id_raw.strip()), "")
        first_note = next((o.note for o in orders if o.note.strip()), "")
        first_title = next((o.title for o in orders if o.title.strip()), "")
        first_tax_id = next((o.tax_id for o in orders if o.tax_id.strip()), "")
        return [ParsedOrder(
            amount_raw=", ".join(amounts),
            order_id_raw=first_oid,
            note=first_note,
            title=first_title,
            tax_id=first_tax_id,
        )]


# ──────────────────────────────────────────────
# 内部解析器
# ──────────────────────────────────────────────


class _KeyValueCellExtractor(HTMLParser):
    """从键值对表格中提取 (key, value) 对（2 列的表格行）"""

    def __init__(self):
        super().__init__()
        self._pairs: list[tuple[str, str]] = []
        self._in_tr = False
        self._cells: list[str] = []
        self._current_cell: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._in_tr = True
            self._cells = []
            self._current_cell = []
        elif tag in ("td", "th") and self._in_tr:
            self._current_cell = []

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._in_tr:
            text = "".join(self._current_cell).strip()
            self._cells.append(text)
            self._current_cell = []
        elif tag == "tr" and self._in_tr:
            if len(self._cells) == 2:
                key = self._cells[0].strip()
                value = self._cells[1].strip()
                self._pairs.append((key, value))
            self._in_tr = False
            self._cells = []
            self._current_cell = []

    def handle_data(self, data):
        if self._in_tr:
            self._current_cell.append(data)

    def handle_entityref(self, name):
        if self._in_tr:
            self._current_cell.append(f"&{name};")

    def handle_charref(self, name):
        if self._in_tr:
            self._current_cell.append(f"&#{name};")

    def get_pairs(self) -> list[tuple[str, str]]:
        return self._pairs


class _WideTableCellExtractor(HTMLParser):
    """从宽表 HTML 中提取二维单元格矩阵"""

    def __init__(self):
        super().__init__()
        self._rows: list[list[str]] = []
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._current_row = []
            self._current_cell = None
        elif tag in ("td", "th") and self._current_row is not None:
            self._current_cell = []

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._current_cell is not None:
            text = "".join(self._current_cell).strip()
            self._current_row.append(text)
            self._current_cell = None
        elif tag == "tr" and self._current_row is not None:
            self._rows.append(self._current_row)
            self._current_row = None
            self._current_cell = None

    def handle_data(self, data):
        if self._current_cell is not None:
            self._current_cell.append(data)

    def handle_entityref(self, name):
        if self._current_cell is not None:
            self._current_cell.append(f"&{name};")

    def handle_charref(self, name):
        if self._current_cell is not None:
            self._current_cell.append(f"&#{name};")

    def get_rows(self) -> list[list[str]]:
        return self._rows
