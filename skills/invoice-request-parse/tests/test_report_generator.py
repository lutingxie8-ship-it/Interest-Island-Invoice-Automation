"""回归测试：report_generator（优化 E — 解析失败单独成段）。

覆盖：
- ReportData.failed_entries / failed_count 字段
- _generate_md 生成 `## 解析失败` 段 + 汇总统计
- _generate_xlsx 生成「解析失败」Sheet 且内容正确
"""
import json
from pathlib import Path

from skill.src.report_generator import ReportData, generate_report


def _make_data(failed_entries=None):
    fe = failed_entries or []
    return ReportData(
        run_time="2026-08-04 11:00:00",
        total_unread=1,
        invoice_count=0,
        urgent_count=0,
        uncertain_count=0,
        failed_count=len(fe),
        urgent_orders=[],
        normal_orders=[],
        invalid_orders=[],
        uncertain_entries=[],
        failed_entries=fe,
    )


def test_md_has_failed_section(tmp_path):
    """回归点 E：.md 报告含独立的『解析失败』段。"""
    fe = [{
        "subject": "s1", "sender": "u1", "date": "2026-08-04",
        "message_id": "m1", "reason": "xlsx 解析失败（未找到开票申请汇总表）",
        "attachment": "a.xlsx",
    }]
    data = _make_data(fe)
    md, xlsx = generate_report(data, str(tmp_path))

    content = Path(md).read_text(encoding="utf-8")
    assert "## 解析失败" in content
    assert "xlsx 解析失败（未找到开票申请汇总表）" in content
    assert "解析失败：**1**" in content  # 汇总统计行


def test_xlsx_has_failed_sheet(tmp_path):
    """回归点 E：.xlsx 报告含『解析失败』Sheet。"""
    fe = [{
        "subject": "s1", "sender": "u1", "date": "2026-08-04",
        "message_id": "m1", "reason": "xlsx 解析失败", "attachment": "a.xlsx",
    }]
    data = _make_data(fe)
    md, xlsx = generate_report(data, str(tmp_path))

    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    assert "解析失败" in wb.sheetnames
    ws = wb["解析失败"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("邮件主题", "发件人", "邮件时间", "失败原因")
    assert rows[1][0] == "s1"
    assert rows[1][3] == "xlsx 解析失败"


def test_report_generates_when_only_failed(tmp_path):
    """回归点 E：即使无成功订单、仅有失败记录，也应生成报告（让失败段可见）。"""
    fe = [{"subject": "s2", "sender": "u2", "date": "2026-08-04", "message_id": "m2", "reason": "附件缺失", "attachment": ""}]
    data = _make_data(fe)
    md, xlsx = generate_report(data, str(tmp_path))
    assert Path(md).exists() and Path(xlsx).exists()
    assert "## 解析失败" in Path(md).read_text(encoding="utf-8")
