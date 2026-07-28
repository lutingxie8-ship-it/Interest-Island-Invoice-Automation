#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
读取税务局导出的Excel发票记录，输出TSV格式（用于粘贴到企微表格）。

用法：
    python read_excel_to_tsv.py <excel文件路径>

输出：TSV格式的发票数据到stdout，每行一条记录。
      统计信息到stderr。

字段映射（Excel「信息汇总表」→ 企微表格）：
    企微开票日期   ← Excel col9  开票日期   (格式化 YYYY/M/D)
    企微发票代码   ← 留空
    企微发票号码   ← Excel col4  数电发票号码
    企微发票类型   ← Excel col22 发票票种
    企微开票名称   ← Excel col8  购买方名称
    企微纳税人识别号 ← Excel col7  购方识别号  (None→空)
    企微开票金额   ← Excel col20 价税合计
    企微订单ID     ← Excel col27 备注
"""
import sys
import openpyxl
from datetime import datetime


def format_date(val):
    """格式化日期：2026-07-24 18:22:02 → 2026/7/24"""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return f"{val.year}/{val.month}/{val.day}"
    s = str(val).strip()
    if not s:
        return ""
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y/%m/%d %H:%M:%S"]:
        try:
            dt = datetime.strptime(s, fmt)
            return f"{dt.year}/{dt.month}/{dt.day}"
        except ValueError:
            continue
    return s


def cell_str(val):
    """转字符串，None→空"""
    if val is None:
        return ""
    # 浮点数如果是整数则去掉小数部分（3280.0 → 3280）
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    s = str(val)
    # 去掉可能的尾部换行（会破坏TSV格式）
    s = s.replace("\n", " ").replace("\r", " ").strip()
    return s


def main():
    if len(sys.argv) < 2:
        print("用法: python read_excel_to_tsv.py <excel文件路径>", file=sys.stderr)
        sys.exit(1)

    filepath = sys.argv[1]
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"错误: 无法打开Excel文件: {e}", file=sys.stderr)
        sys.exit(1)

    # 优先用「信息汇总表」，没有则用第一个sheet
    sheet_name = None
    for sn in wb.sheetnames:
        if '信息汇总' in sn:
            sheet_name = sn
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
        print(f"提示: 未找到'信息汇总表'，使用'{sheet_name}'", file=sys.stderr)

    ws = wb[sheet_name]

    # Excel列索引（1-based）→ row数组索引（0-based）的映射：
    # col 9 → row[8]  开票日期
    # col 4 → row[3]  数电发票号码
    # col 22 → row[21] 发票票种
    # col 8 → row[7]  购买方名称
    # col 7 → row[6]  购方识别号
    # col 20 → row[19] 价税合计
    # col 27 → row[26] 备注
    # 企微TSV顺序: 开票日期, (空), 发票号码, 发票类型, 开票名称, 纳税人识别号, 开票金额, 订单ID

    lines = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # 跳过空行
        if not row or all(v is None for v in row):
            skipped += 1
            continue

        # 跳过合计行（第一列是"合计行"或类似）
        first_cell = str(row[0]) if row[0] else ""
        if "合计" in first_cell:
            skipped += 1
            continue

        # 确保有足够列
        if len(row) < 27:
            skipped += 1
            continue

        # 提取字段
        date = format_date(row[8])        # col 9 开票日期
        invoice_code = ""                 # 企微发票代码列留空
        invoice_num = cell_str(row[3])    # col 4 数电发票号码
        invoice_type = cell_str(row[21])  # col 22 发票票种
        name = cell_str(row[7])           # col 8 购买方名称
        tax_id = cell_str(row[6])         # col 7 购方识别号
        amount = cell_str(row[19])        # col 20 价税合计
        order_id = cell_str(row[26])      # col 27 备注

        # 跳过没有发票号码的行
        if not invoice_num:
            skipped += 1
            continue

        # 组装TSV行（制表符分隔，发票代码列为空→两个连续\t）
        tsv_line = "\t".join([
            date,          # 开票日期
            invoice_code,  # 发票代码（空）
            invoice_num,   # 发票号码
            invoice_type,  # 发票类型
            name,          # 开票名称
            tax_id,        # 纳税人识别号
            amount,        # 开票金额
            order_id,      # 订单ID
        ])
        lines.append(tsv_line)

    wb.close()

    # 输出TSV到stdout
    print("\n".join(lines))

    # 统计信息到stderr
    print(f"\n# 共 {len(lines)} 条记录，跳过 {skipped} 行", file=sys.stderr)


if __name__ == "__main__":
    main()
